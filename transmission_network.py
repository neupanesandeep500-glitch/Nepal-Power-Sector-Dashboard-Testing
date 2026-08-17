"""
transmission_network.py — Nepal Transmission Network Map module
==================================================================
Everything behind the "🔌 Transmission Network" tab lives here, kept
independent of app.py / data_engine.py (the DoED licensing pipeline)
and of NEA.py (NEA operational performance), exactly the way those two
are already kept independent of each other in this codebase.

WHAT THIS FILE DOES
--------------------
1. WORKBOOK PARSING — reads the two-sheet "Nepal Transmission Network —
   Editable Data Sheet" workbook (Substations / Transmission Lines,
   see the workbook's own Data Dictionary tab for the exact columns)
   and turns it into the JSON the Leaflet map + Chart.js distribution
   panel consume.
2. LINE GEOMETRY — the editable workbook intentionally does NOT carry
   the original shapefile-derived route polylines (those live only in
   the source .shp files). Instead each Transmission Lines row names a
   From/To substation, matched by exact name back to the Substations
   tab. So every line is drawn here as a straight segment between its
   two resolved substation coordinates — this is what makes the sheet
   editable/re-importable in the first place (see the workbook's
   "Do not edit or reuse [ID]" note). Rows whose From/To can't be
   resolved are still listed in the data table/detail panel but are
   skipped on the map itself.
3. LIVE SYNC / ADMIN UPLOAD — same pattern as NEA.py: an optional
   Google Sheet URL (public CSV/xlsx export, no auth) OR a manual
   .xlsx upload via the admin panel, cached to disk, refreshed on a
   background timer, with a bundled-snapshot fallback so the tab is
   never blank on first boot.
4. DISTRIBUTION STATS — pre-aggregates counts/length-km by Stage ×
   Voltage and Stage × Province server-side, for the stacked bar
   panel (stacked by Stage, per the admin's request), so the template
   stays a dumb renderer.

HOW TO WIRE THIS INTO app.py
-----------------------------
    import transmission_network as tn

    tn.bootstrap()                 # call once at startup (non-blocking)
    tn.start_background_refresh()  # call once at startup

    @server.route("/transmission-network-map")
    def serve_tn_map():
        return tn.render_map_html()

    # inside render_tab(), for tab == "transmission_network":
    #     html.Iframe(src="/transmission-network-map", ...)
"""

from __future__ import annotations

import json
import math
import os
import re
import threading
import traceback
from datetime import datetime

import openpyxl

_HERE = os.path.dirname(os.path.abspath(__file__))

CACHE_WORKBOOK_PATH = os.path.join(_HERE, "transmission_network_workbook_cache.xlsx")
TEMPLATE_PATH = os.path.join(_HERE, "nea_assets", "transmission_network_template.html")
FALLBACK_WORKBOOK_PATH = os.path.join(_HERE, "nea_assets", "transmission_network_fallback.xlsx")

AUTO_REFRESH_HOURS = float(os.environ.get("TN_AUTO_REFRESH_HOURS", "6"))

_TN_CONFIG_PATH = os.path.join(os.environ.get("DATA_DIR", _HERE), "transmission_network_config.json")

STAGE_ORDER = ["Existing", "Under Construction", "Future", "Proposed", "Reference"]
STAGE_COLORS = {
    "Existing": "#3FC7C0",
    "Under Construction": "#F2B84B",
    "Future": "#F0576B",
    "Proposed": "#8C7CF0",
    "Reference": "#6B7690",
}
VOLTAGE_COLORS = {132: "#F2B84B", 220: "#3FC7C0", 400: "#F0576B", 66: "#8C7CF0", 33: "#6B7690"}


# ── Sheet URL persistence (mirrors NEA.py) ──────────────────────────────

def _load_persisted_sheet_url():
    try:
        with open(_TN_CONFIG_PATH, "r", encoding="utf-8") as f:
            return (json.load(f).get("sheet_url") or "").strip() or None
    except Exception:
        return None


def _save_persisted_sheet_url(url):
    try:
        os.makedirs(os.path.dirname(_TN_CONFIG_PATH) or ".", exist_ok=True)
        with open(_TN_CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump({"sheet_url": url}, f)
    except Exception:
        traceback.print_exc()


def _resolve_sheet_url():
    return _load_persisted_sheet_url() or os.environ.get("TN_SHEET_URL") or None


def current_sheet_url() -> str:
    return _resolve_sheet_url() or ""


def has_persisted_sheet_url() -> bool:
    return _load_persisted_sheet_url() is not None


def clear_persisted_sheet_url() -> bool:
    try:
        if os.path.exists(_TN_CONFIG_PATH):
            os.remove(_TN_CONFIG_PATH)
    except Exception:
        traceback.print_exc()
    url = _resolve_sheet_url()
    return refresh(url) if url else False


# ── Parsing helpers ──────────────────────────────────────────────────────

def _norm(s):
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()


def _s(v):
    if v is None:
        return None
    v = str(v).strip()
    return v or None


def _f(v):
    try:
        if v is None or str(v).strip() == "":
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _header_map(ws):
    row1 = [c.value for c in ws[1]]
    return {_norm(h): i for i, h in enumerate(row1) if h}


def _parse_substations(ws):
    hm = _header_map(ws)

    def col(*names):
        for n in names:
            if n in hm:
                return hm[n]
        return None

    c_id, c_name, c_stage = col("id"), col("name"), col("stage")
    c_lat, c_lon = col("latitude"), col("longitude")
    c_prov, c_dist, c_local = col("province"), col("district"), col("local body")
    c_volt = col("voltage level (kv)")
    c_cap = col("transformer capacity")
    c_load = col("connected load / feeders")
    c_notes = col("notes")

    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        sid = _s(row[c_id]) if c_id is not None else None
        name = _s(row[c_name]) if c_name is not None else None
        if not sid or not name:
            continue
        out.append({
            "id": sid,
            "name": name,
            "stage": _s(row[c_stage]) if c_stage is not None else "Existing",
            "lat": _f(row[c_lat]) if c_lat is not None else None,
            "lon": _f(row[c_lon]) if c_lon is not None else None,
            "province": _s(row[c_prov]) if c_prov is not None else None,
            "district": _s(row[c_dist]) if c_dist is not None else None,
            "local_body": _s(row[c_local]) if c_local is not None else None,
            "voltage_level": _s(row[c_volt]) if c_volt is not None else None,
            "capacity": _s(row[c_cap]) if c_cap is not None else None,
            "load": _s(row[c_load]) if c_load is not None else None,
            "notes": _s(row[c_notes]) if c_notes is not None else None,
        })
    return out


def _parse_lines(ws, subs_by_name):
    hm = _header_map(ws)

    def col(*names):
        for n in names:
            if n in hm:
                return hm[n]
        return None

    c_id, c_label = col("id"), col("label")
    c_from, c_to = col("from substation"), col("to substation")
    c_volt, c_stage = col("voltage (kv)"), col("stage")
    c_year = col("expected completion year")
    c_circuit, c_conductor = col("circuit"), col("conductor")
    c_len = col("length (km)")
    c_prov, c_dist = col("province(s)"), col("district(s)")
    c_class = col("line classification")
    c_notes = col("notes")

    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        lid = _s(row[c_id]) if c_id is not None else None
        if not lid:
            continue
        from_name = _s(row[c_from]) if c_from is not None else None
        to_name = _s(row[c_to]) if c_to is not None else None
        from_sub = subs_by_name.get(_norm(from_name)) if from_name else None
        to_sub = subs_by_name.get(_norm(to_name)) if to_name else None
        from_coord = [from_sub["lat"], from_sub["lon"]] if from_sub and from_sub["lat"] and from_sub["lon"] else None
        to_coord = [to_sub["lat"], to_sub["lon"]] if to_sub and to_sub["lat"] and to_sub["lon"] else None

        provinces = [p.strip() for p in re.split(r"[,;]", str(row[c_prov] or "")) if p.strip()] if c_prov is not None else []
        districts = [d.strip() for d in re.split(r"[,;]", str(row[c_dist] or "")) if d.strip()] if c_dist is not None else []

        out.append({
            "id": lid,
            "label": _s(row[c_label]) if c_label is not None else lid,
            "from": from_name,
            "to": to_name,
            "voltage_kv": _f(row[c_volt]) if c_volt is not None else None,
            "stage": _s(row[c_stage]) if c_stage is not None else "Existing",
            "expected_completion": _s(row[c_year]) if c_year is not None else None,
            "circuit": _s(row[c_circuit]) if c_circuit is not None else None,
            "conductor": _s(row[c_conductor]) if c_conductor is not None else None,
            "length_km": _f(row[c_len]) if c_len is not None else None,
            "provinces": provinces,
            "districts": districts,
            "classification": _s(row[c_class]) if c_class is not None else None,
            "notes": _s(row[c_notes]) if c_notes is not None else None,
            "coords": [from_coord, to_coord] if (from_coord and to_coord) else None,
            "resolved": bool(from_coord and to_coord),
        })
    return out


def _find_sheet(wb, *aliases):
    by_norm = {_norm(n): n for n in wb.sheetnames}
    for a in aliases:
        if _norm(a) in by_norm:
            return wb[by_norm[_norm(a)]]
    raise ValueError(f"Workbook is missing a sheet matching any of: {aliases} "
                      f"(found: {wb.sheetnames})")


def parse_workbook(path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws_sub = _find_sheet(wb, "substations")
    ws_line = _find_sheet(wb, "transmission lines")

    substations = _parse_substations(ws_sub)
    subs_by_name = {_norm(s["name"]): s for s in substations if s["name"]}
    lines = _parse_lines(ws_line, subs_by_name)

    return {"substations": substations, "lines": lines}


def _stage_sort_key(stage):
    try:
        return STAGE_ORDER.index(stage)
    except ValueError:
        return len(STAGE_ORDER)


def build_distribution_stats(parsed: dict) -> dict:
    """Pre-aggregate counts + route-km, stacked by Stage, sliced by
    Voltage and by Province — feeds the distribution panel's Chart.js
    stacked bar (toggle between the two slices client-side)."""
    lines = parsed["lines"]
    subs = parsed["substations"]

    stages_present = sorted({l["stage"] or "Unknown" for l in lines} | {s["stage"] or "Unknown" for s in subs},
                             key=_stage_sort_key)

    # By voltage (lines only — km is the meaningful unit here)
    voltages_present = sorted({int(l["voltage_kv"]) for l in lines if l["voltage_kv"]})
    by_voltage = {"categories": [f"{v} kV" for v in voltages_present], "stages": stages_present,
                  "count": {}, "km": {}}
    for stage in stages_present:
        by_voltage["count"][stage] = [
            sum(1 for l in lines if (l["stage"] or "Unknown") == stage and l["voltage_kv"] == v)
            for v in voltages_present
        ]
        by_voltage["km"][stage] = [
            round(sum((l["length_km"] or 0) for l in lines
                      if (l["stage"] or "Unknown") == stage and l["voltage_kv"] == v), 1)
            for v in voltages_present
        ]

    # By province (lines — a line can touch multiple provinces, counted once per province it touches)
    provinces_present = sorted({p for l in lines for p in (l["provinces"] or [])})
    by_province = {"categories": provinces_present, "stages": stages_present, "count": {}, "km": {}}
    for stage in stages_present:
        by_province["count"][stage] = [
            sum(1 for l in lines if (l["stage"] or "Unknown") == stage and p in (l["provinces"] or []))
            for p in provinces_present
        ]
        by_province["km"][stage] = [
            round(sum((l["length_km"] or 0) for l in lines
                      if (l["stage"] or "Unknown") == stage and p in (l["provinces"] or [])), 1)
            for p in provinces_present
        ]

    # Substation counts by stage (small summary card, not stacked)
    substations_by_stage = {stage: sum(1 for s in subs if (s["stage"] or "Unknown") == stage)
                             for stage in stages_present}

    kpi = {
        "total_lines": len(lines),
        "total_km": round(sum((l["length_km"] or 0) for l in lines), 1),
        "total_substations": len(subs),
        "unresolved_lines": sum(1 for l in lines if not l["resolved"]),
    }

    return {
        "stages": stages_present,
        "stage_colors": STAGE_COLORS,
        "by_voltage": by_voltage,
        "by_province": by_province,
        "substations_by_stage": substations_by_stage,
        "kpi": kpi,
    }


def build_map_data(parsed: dict) -> dict:
    return {
        "substations": parsed["substations"],
        "lines": parsed["lines"],
        "stats": build_distribution_stats(parsed),
        "stage_colors": STAGE_COLORS,
        "voltage_colors": {str(k): v for k, v in VOLTAGE_COLORS.items()},
    }


# ── Live sync / cache (mirrors NEA.py) ──────────────────────────────────

_CACHE = {"data": None, "parsed": None, "last_sync": None, "source": None, "error": None}
_lock = threading.Lock()


def _load_bundled_fallback() -> dict:
    if os.path.exists(FALLBACK_WORKBOOK_PATH):
        return build_map_data(parse_workbook(FALLBACK_WORKBOOK_PATH))
    return {}


def _download_google_sheet_xlsx(url_or_id, out_path):
    import urllib.request
    import urllib.error

    sheet_id = url_or_id
    if "/" in url_or_id:
        for pat in (r"/spreadsheets/d/([a-zA-Z0-9-_]+)", r"/file/d/([a-zA-Z0-9-_]+)", r"id=([a-zA-Z0-9-_]+)"):
            m = re.search(pat, url_or_id)
            if m:
                sheet_id = m.group(1)
                break

    candidate_urls = [
        f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx",
        f"https://drive.google.com/uc?export=download&id={sheet_id}",
    ]
    last_err = None
    for export_url in candidate_urls:
        try:
            req = urllib.request.Request(export_url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=120) as resp:
                content = resp.read()
                ctype = resp.headers.get("Content-Type", "")
        except urllib.error.HTTPError as e:
            last_err = f"HTTP {e.code} from {export_url}"
            continue
        except Exception as e:
            last_err = f"{type(e).__name__}: {e} ({export_url})"
            continue

        if content[:2] != b"PK":
            snippet = content[:200].decode("utf-8", errors="replace")
            last_err = f"{export_url} did not return a valid .xlsx (Content-Type={ctype!r}; started with: {snippet!r})"
            continue

        with open(out_path, "wb") as f:
            f.write(content)
        return out_path

    raise RuntimeError(
        f"Could not download the Transmission Network workbook (id={sheet_id}). Last error: {last_err}. "
        f"Confirm the sheet is shared as 'Anyone with the link can view'."
    )


def refresh(sheet_url: str = None) -> bool:
    sheet_url = sheet_url or _resolve_sheet_url()
    if not sheet_url:
        with _lock:
            if _CACHE["data"] is None:
                fallback = _load_bundled_fallback()
                if fallback:
                    _CACHE.update(data=fallback, source="Bundled fallback snapshot", error=None)
        return _CACHE["data"] is not None

    with _lock:
        try:
            _download_google_sheet_xlsx(sheet_url, CACHE_WORKBOOK_PATH)
            parsed = parse_workbook(CACHE_WORKBOOK_PATH)
            data = build_map_data(parsed)
            _CACHE.update(data=data, parsed=parsed,
                          last_sync=datetime.now().strftime("%Y-%m-%d %H:%M"),
                          source="Google Sheet (live sync)", error=None)
            return True
        except Exception as exc:
            traceback.print_exc()
            _CACHE["error"] = str(exc)
            if _CACHE["data"] is None and os.path.exists(CACHE_WORKBOOK_PATH):
                try:
                    parsed = parse_workbook(CACHE_WORKBOOK_PATH)
                    _CACHE.update(data=build_map_data(parsed), parsed=parsed,
                                  source="Cached workbook (last good sync)")
                except Exception:
                    traceback.print_exc()
            if _CACHE["data"] is None:
                fallback = _load_bundled_fallback()
                if fallback:
                    _CACHE["data"] = fallback
                    _CACHE["source"] = "Bundled fallback snapshot"
            return False


def set_sheet_url(url: str) -> bool:
    url = (url or "").strip()
    if not url:
        raise ValueError("Please provide a Google Sheet URL or ID for the Transmission Network data")
    _save_persisted_sheet_url(url)
    return refresh(url)


def load_from_path(path: str) -> bool:
    """Load from an admin-uploaded .xlsx. Copies to the cache path so it
    persists across restarts. Returns True on success."""
    with _lock:
        try:
            parsed = parse_workbook(path)
            data = build_map_data(parsed)
            import shutil
            shutil.copy(path, CACHE_WORKBOOK_PATH)
            _CACHE.update(data=data, parsed=parsed,
                          last_sync=datetime.now().strftime("%Y-%m-%d %H:%M"),
                          source="Admin upload", error=None)
            return True
        except Exception as exc:
            traceback.print_exc()
            _CACHE["error"] = str(exc)
            return False


def bootstrap():
    persisted_url = _load_persisted_sheet_url()
    env_url = os.environ.get("TN_SHEET_URL")
    if persisted_url or env_url:
        print(f"[TN DEBUG] Transmission Network sheet URL configured — attempting sync.")
    elif os.path.exists(CACHE_WORKBOOK_PATH):
        print("[TN DEBUG] No sheet URL configured; using last cached workbook on disk.")
    elif os.path.exists(FALLBACK_WORKBOOK_PATH):
        print("[TN DEBUG] No sheet URL configured; using bundled fallback workbook.")
    else:
        print("[TN DEBUG] No Transmission Network data source configured yet — "
              "upload a workbook at /admin to populate the map.")

    def _bootstrap_and_report():
        # Prefer the last cached upload/sync over a (possibly unset) sheet URL.
        if not _resolve_sheet_url() and os.path.exists(CACHE_WORKBOOK_PATH):
            try:
                parsed = parse_workbook(CACHE_WORKBOOK_PATH)
                with _lock:
                    _CACHE.update(data=build_map_data(parsed), parsed=parsed,
                                  source="Cached workbook (last good sync)")
                print("[TN DEBUG] Loaded last cached workbook from disk.")
                return
            except Exception:
                traceback.print_exc()
        ok = refresh()
        status = sync_status()
        if ok:
            print(f"[TN DEBUG] Initial load succeeded: {status['source']}")
        else:
            print(f"[TN DEBUG] Initial load failed: {status['error']}")

    threading.Thread(target=_bootstrap_and_report, daemon=True).start()


def start_background_refresh():
    interval = max(AUTO_REFRESH_HOURS, 0.25) * 3600

    def _tick():
        if _resolve_sheet_url():
            refresh()
        t = threading.Timer(interval, _tick)
        t.daemon = True
        t.start()

    t = threading.Timer(interval, _tick)
    t.daemon = True
    t.start()


def get_map_data() -> dict:
    return _CACHE["data"] or _load_bundled_fallback() or {
        "substations": [], "lines": [],
        "stats": {"stages": [], "stage_colors": STAGE_COLORS,
                  "by_voltage": {"categories": [], "stages": [], "count": {}, "km": {}},
                  "by_province": {"categories": [], "stages": [], "count": {}, "km": {}},
                  "substations_by_stage": {}, "kpi": {"total_lines": 0, "total_km": 0,
                                                        "total_substations": 0, "unresolved_lines": 0}},
        "stage_colors": STAGE_COLORS, "voltage_colors": {},
    }


def sync_status() -> dict:
    return {"last_sync": _CACHE["last_sync"], "source": _CACHE["source"], "error": _CACHE["error"]}


def get_admin_status() -> dict:
    with _lock:
        return {
            "last_sync": _CACHE["last_sync"],
            "source": _CACHE["source"],
            "error": _CACHE["error"],
            "has_data": _CACHE["data"] is not None,
        }


# ── HTML rendering ────────────────────────────────────────────────────

def render_map_html() -> str:
    with open(TEMPLATE_PATH, "r", encoding="utf-8") as f:
        template = f.read()
    data_json = json.dumps(get_map_data())
    return template.replace("__TN_DATA_JSON__", data_json)


# ── Smoke test ──────────────────────────────────────────────────────

if __name__ == "__main__":
    ok = refresh()
    print("initial load ok:", ok, sync_status())
    d = get_map_data()
    print("substations:", len(d["substations"]), "lines:", len(d["lines"]))
    print("kpi:", d["stats"]["kpi"])
